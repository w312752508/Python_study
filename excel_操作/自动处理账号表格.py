from openpyxl import load_workbook
import xlwings as xw
def del_(delete_list,filename,sheet_name):
    app = xw.App(visible=False,add_book=False) #操作时不显示excel表格，不新建表格
    app.display_alerts = False
    app.screen_updating = False
    wb = app.books.open(filename) #打开excel表格
    sheet = wb.sheets(sheet_name) #实例化工作簿
    temp_del = 0  #每次删除行号会变动，设置临时值，每次删除自动运算新行号
    if len(delete_list) > 0:
        for delete_row in delete_list:
            sheet.range('A' + str(delete_row - temp_del)).api.EntireRow.Delete()
            temp_del = temp_del + 1
    wb.save()
    wb.close()

def update_excel():
    file_dir = "d:\\excel_test\\user_list.xlsx"
    wb= load_workbook(file_dir)
    wb_sheet1 =wb["堡垒机"]
    wb_sheet2 =wb["跳板机"]
    with open("update_list.txt","r") as fs :
        for i in fs :
            num_list1 =[]
            num_list2 =[]
            i_list = i.split(",")
            for r in wb_sheet1.rows:
                for ce in r :
                    if str(i_list[0].strip()) == str(ce.value) :
                        num_list1.append(True)
                        wb_sheet1.cell(row=ce.row, column=9).value = i_list[1].strip()
                        wb_sheet1.cell(row=ce.row, column=10).value = i_list[2].strip()
                        wb_sheet1.cell(row=ce.row, column=11).value = i_list[3].strip()
            if True not in num_list1:
                print(i_list[0],"堡垒机账号不存在")
            for r2 in wb_sheet2.rows:
                for ce2 in r2:
                    if str(i_list[0].strip()) == str(ce2.value):
                        num_list2.append(True)
                        print(wb_sheet2.cell(row=ce2.row, column=1).value,i_list[0].strip(),
                              wb_sheet2.cell(row=ce2.row, column=6).value,"已更新账号有效期至",i_list[3].strip())
                        wb_sheet2.cell(row=ce2.row, column=9).value = i_list[1].strip()
                        wb_sheet2.cell(row=ce2.row, column=10).value = i_list[2].strip()
                        wb_sheet2.cell(row=ce2.row, column=11).value = i_list[3].strip()
            if True not in num_list2:
                print(i_list[0],"跳板机账号不存在")

    wb.save(file_dir)

def del_excel():
    file_dir = "d:\\excel_test\\user_list.xlsx"
    b_delete_list = []
    t_delete_list = []
    del_record_list=[]
    # with open("del_list.txt","r") as fs :
    #     wb = load_workbook(file_dir)
    #     wb_sheet1 = wb["堡垒机"]
    #     for i in fs :
    #         num_list1 =[]  #定义账号是否存在空列表
    #         num_line = 0  #存储行号的空列表
    #         for r in wb_sheet1.rows:
    #             num_line += 1
    #             if r[0].value == str(i.strip()):
    #                 num_list1.append(True)
    #                 del_record_list.append(r[0])
    #                 b_delete_list.append(num_line)
    #         if True not in num_list1:
    #             print(str(i.strip()),"堡垒机账号不存在")
    #     wb.close()
    #     b_delete_list = sorted(b_delete_list) #行号排序，如果行号无序，后面运算会出错
    #     del_(b_delete_list, file_dir,"堡垒机")
    # with open("del_list.txt", "r") as fs:
    #     wb = load_workbook(file_dir)
    #     wb_sheet2 = wb["跳板机"]
    #     for i in fs :
    #         num_list2 =[]
    #         num_line2 = 0
    #         for r2 in wb_sheet2.rows:
    #             num_line2 += 1
    #             if r2[3].value == str(i.strip()):
    #                 num_list2.append(True)
    #                 del_record_list.append(r[0])
    #                 t_delete_list.append(num_line2)
    #         if True not in num_list2:
    #             print(str(i.strip()), "跳板机账号不存在")
    #     wb.close()
    #     t_delete_list=sorted(t_delete_list)
    #     del_(t_delete_list, file_dir,"跳板机")

    del_record_list =[["dpzh021389","赵海"],["dpzh02","赵海"]]
    # del_record_list = set(del_record_list)
    def inset_record(file_dir,name_list):
        wb = load_workbook(file_dir)
        wb_sheet = wb["账号清理记录"]
        for r in wb_sheet.rows:
            if len(name_list) > 0:
                if len(str(r[0].value)) == 0:
                    wb_sheet.cell(row=r.row, column=0).value = name_list[0][0]
                    wb_sheet.cell(row=r.row, column=2).value = name_list[0][1]
                    del name_list[0]
    inset_record(file_dir,del_record_list)

user_select = input("1、更新用户信息\n"
                    "2、删除用户信息\n"
                    "选择表格操作类型：")
if user_select == "1":
    update_excel()
elif user_select == "2":
    del_excel()
