from openpyxl import load_workbook
import xlwings as xw
from org_conf import org_dic
def add_user(user_name):
    file_dir = "d:\\excel_test\\user_list.xlsx"
    wb = load_workbook(file_dir)
    wb_sheet1 = wb["堡垒机"]
    wb_sheet2 = wb["跳板机"]
    wb_sheet3 = wb["win系统账号"]
    wb_sheet4 = wb["人员基本信息"]
    name_num = user_name[0][-6:]
    for r in wb_sheet4.rows:
        if r[0].value == name_num:
            name = r[1].value
            name_org = r[2].value
            break
        else:
            name_org = False
    if name_org == False:
        print(name_num,"查无此人,无添加记录")
        exit()
    else:
        for k in org_dic:
            for k1 in org_dic[k]:
                if name_org in org_dic[k][k1]:
                    name_up_org=k1
        req_num =user_name[1]
        if len(user_name) > 2:
            b_add_list = [user_name[0],name_num,name,name_org,name_up_org,"IT","普通用户","活动",req_num,user_name[2],user_name[3]]
            t_add_list = ["","","",user_name[0],name_num,name,name_org,name_up_org,req_num,user_name[2],user_name[3]]
            s_add_list = ["","",user_name[0],"",name,req_num,user_name[2],user_name[3]]
        else:
            b_add_list = [user_name[0], name_num, name, name_org, name_up_org, "IT", "普通用户", "活动", req_num]
            t_add_list = ["", "", "", user_name[0], name_num, name, name_org, name_up_org, req_num]
            s_add_list = ["", "", user_name[0], "", name, req_num]
        wb_sheet1.append(b_add_list)
        print(user_name[0],"堡垒机信息添加完成")
        wb_sheet2.append(t_add_list)
        print(user_name[0], "跳板机机信息添加完成")
        # wb_sheet3.append(s_add_list)
        # print(user_name[0], "系统信息添加完成")
        wb.save(file_dir)

def del_(delete_list,filename,sheet_name): #定义删除行的函数
    app = xw.App(visible=False,add_book=False) #操作时不显示excel表格，不新建表格
    app.display_alerts = False
    app.screen_updating = False
    wb = app.books.open(filename) #打开excel表格
    sheet = wb.sheets(sheet_name) #实例化工作簿
    temp_del = 0  #每次删除行号会变动，设置临时值，每次删除自动运算新行号
    if len(delete_list) > 0:
        for delete_row in delete_list:
            sheet.range('A' + str(delete_row - temp_del)).api.EntireRow.Delete()
            temp_del = temp_del + 1
    wb.save()
    wb.close()

def inset_record(file_dir,del_record_list): #用户删除信息登记函数
    wb = load_workbook(file_dir)
    wb_sheet = wb["账号清理记录"]
    if len(del_record_list) == 0:
        print("无删除账号，无删除记录需添加")
    else:
        for i in del_record_list:
            i += ["", "离职"]
            wb_sheet.append(i)
            print(i[0], "删除记录添加成功")
        wb.save(file_dir)

def update_excel(): #用户信息更新函数
    file_dir = "d:\\excel_test\\user_list.xlsx"
    wb= load_workbook(file_dir)
    wb_sheet1 =wb["堡垒机"]
    wb_sheet2 =wb["跳板机"]
    with open("update_list.txt","r") as fs :
        for i in fs :
            num_list1 =[]
            num_list2 =[]
            i_list = i.split(",")
            for r in wb_sheet1.rows:
                for ce in r :
                    if str(i_list[0].strip()) == str(ce.value) :
                        num_list1.append(True)
                        wb_sheet1.cell(row=ce.row, column=9).value = i_list[1].strip()
                        wb_sheet1.cell(row=ce.row, column=10).value = i_list[2].strip()
                        wb_sheet1.cell(row=ce.row, column=11).value = i_list[3].strip()
            if True not in num_list1:
                print(i_list[0],"堡垒机账号不存在")
            for r2 in wb_sheet2.rows:
                for ce2 in r2:
                    if str(i_list[0].strip()) == str(ce2.value):
                        num_list2.append(True)
                        print(wb_sheet2.cell(row=ce2.row, column=1).value,i_list[0].strip(),
                              wb_sheet2.cell(row=ce2.row, column=6).value,"已更新账号有效期至",i_list[3].strip())
                        wb_sheet2.cell(row=ce2.row, column=9).value = i_list[1].strip()
                        wb_sheet2.cell(row=ce2.row, column=10).value = i_list[2].strip()
                        wb_sheet2.cell(row=ce2.row, column=11).value = i_list[3].strip()
            if True not in num_list2:
                print(i_list[0],"跳板机账号不存在")

    wb.save(file_dir)

def del_excel(): #删除用户
    file_dir = "d:\\excel_test\\user_list.xlsx"
    b_delete_list = []
    t_delete_list = []
    w_delete_list = []
    del_record_list=[]
    with open("del_list.txt","r") as fs :
        wb = load_workbook(file_dir)
        wb_sheet1 = wb["堡垒机"]
        for i in fs :
            num_list1 =[]  #定义账号是否存在空列表
            num_line = 0  #存储行号的空列表
            for r in wb_sheet1.rows:
                num_line += 1
                if r[0].value == str(i.strip()):
                    num_list1.append(True)
                    b_delete_list.append(num_line)
                    if [r[0].value,r[2].value] not in del_record_list:
                        del_record_list.append([r[0].value,r[2].value])
                    print(r[0].value,"堡垒机账号删除成功")
            if True not in num_list1:
                print(str(i.strip()),"堡垒机账号不存在")
        wb.close()
        b_delete_list = sorted(b_delete_list) #行号排序，如果行号无序，后面运算会出错
        del_(b_delete_list, file_dir,"堡垒机")
    with open("del_list.txt", "r") as fs:
        wb = load_workbook(file_dir)
        wb_sheet2 = wb["跳板机"]
        for i in fs :
            num_list2 =[]
            num_line2 = 0
            for r2 in wb_sheet2.rows:
                num_line2 += 1
                if r2[3].value == str(i.strip()):
                    num_list2.append(True)
                    t_delete_list.append(num_line2)
                    if [r2[3].value, r2[5].value] not in del_record_list:
                        del_record_list.append([r2[3].value, r2[5].value])
                    print(r2[3].value,r2[0].value,"跳板机账号删除成功")
            if True not in num_list2:
                print(str(i.strip()), "跳板机账号不存在")
        wb.close()
        t_delete_list=sorted(t_delete_list)
        del_(t_delete_list, file_dir,"跳板机")
    with open("del_list.txt", "r") as fs:
        wb = load_workbook(file_dir)
        wb_sheet2 = wb["win系统账号"]
        for i in fs :
            num_list2 =[]
            num_line2 = 0
            for r3 in wb_sheet2.rows:
                num_line2 += 1
                if r3[2].value == str(i.strip()):
                    num_list2.append(True)
                    w_delete_list.append(num_line2)
                    if [r3[2].value, r3[4].value] not in del_record_list:
                        del_record_list.append([r3[2].value, r3[4].value])
                    print(r3[2].value,r3[1].value,"服务器账号删除成功")
            if True not in num_list2:
                print(str(i.strip()), "服务器账号不存在")
        wb.close()
        w_delete_list=sorted(w_delete_list)
        del_(w_delete_list, file_dir,"win系统账号")
    inset_record(file_dir, del_record_list)

def add_excel():
    with open("add_list.txt","r") as fs :
        for i in fs :
            i = i.strip().split(",")
            add_user(i)
user_select = input("1、新增用户信息\n"
                    "2、更新用户信息\n"
                    "3、删除用户信息\n"
                    "选择表格操作类型：")
if user_select == "1":
    add_excel()
elif user_select == "2":
    update_excel()
elif user_select == "3":
    del_excel()


