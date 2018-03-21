from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.utils import get_column_letter
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
#新建一个workbook工作簿
wb = Workbook()

#第一个sheet
ws = wb.worksheets[0]

#修改第一个工作表名
ws.title = "下单统计2"

#增加一个sheet,index值为工作表位置，按顺序增加，不能跳增
wb.create_sheet(u"结账统计",index=1)

# 删除某个工作表
#wb.remove(sheet)
#del wb[sheet]

# #指定工作表位置添加信息
ws_sheet = wb['结账统计']
ws_sheet["a1"].value = "hello"
#从第一个空白行开始添加一行信息
row = [1 ,2, 3, 4, 5]
# ws_sheet.append(row)
#添加多行信息
rows = [['Number', 'data1', 'data2'],[2, 40, 30],[3, 40, 25],[4, 50, 30],[5, 30, 10],[6, 25, 5]]
for i in rows:
    ws_sheet.append(i)


file_name = 'new_test.xlsx'
file_dir = 'd:\\excel_test\\'
dest_filename = '%s%s'%(file_dir,file_name)
wb.save(dest_filename)

# #给A1赋值
# ws.cell('A1').value = "%s" %("跟随总数")
# #给A2赋值
# #先把数字转换成字母
# col = get_column_letter(1)
# #赋值
# ws.cell('%s%s'%(col, 2)).value = '%s' % ("A2")
# #字体修改样式
# ##颜色
# ws.cell('A2').style.font.color.index =Color.GREEN
# ##字体名称
# ws.cell('A2').style.font.name ='Arial'
# ##字号
# ws.cell('A2').style.font.size =8
# ##加粗
# ws.cell('A2').style.font.bold =True
# ##不知道干啥用的
# ws.cell('A2').style.alignment.wrap_text =True
# ##背景 好像不太好用 是个BUG
# ws.cell('A2').style.fill.fill_type =Fill.FILL_SOLID
# ws.cell('A2').style.fill.start_color.index =Color.DARKRED
# ##修改某一列宽度
# ws.column_dimensions["C"].width =60.0
# ##增加一个表
# ws = wb.create_sheet()
# ws.title = u'结单统计'
# ##保存生成xml
# file_name = 'new_test.xlsx'
# file_dir = 'd:\\excel_test\\'
# dest_filename = '%s%s'%(file_dir,file_name)
# ew = ExcelWriter(workbook = wb)
# ew = ExcelWriter(workbook = wb)