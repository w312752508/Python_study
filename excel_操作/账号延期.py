from openpyxl import load_workbook

file_dir = "d:\\excel_test\\user_list.xlsx"
wb= load_workbook(file_dir)
wb_sheet1 =wb["堡垒机"]
wb_sheet2 =wb["跳板机"]
with open("user_list.txt","r") as fs :
    for i in fs :
        num_list1 =[]
        num_list2 =[]
        i_list = i.split(",")
        for r in wb_sheet1.rows:
            for ce in r :
                if str(i_list[0].strip()) == str(ce.value) :
                    num_list1.append(True)
                    wb_sheet1.cell(row=ce.row, column=9).value = i_list[1].strip()
                    wb_sheet1.cell(row=ce.row, column=10).value = i_list[2].strip()
                    wb_sheet1.cell(row=ce.row, column=11).value = i_list[3].strip()
        if True not in num_list1:
            print(i_list[0],"堡垒机账号不存在")
        for r2 in wb_sheet2.rows:
            for ce2 in r2:
                if str(i_list[0].strip()) == str(ce2.value):
                    num_list2.append(True)
                    print(wb_sheet2.cell(row=ce2.row, column=1).value,i_list[0].strip(),
                          wb_sheet2.cell(row=ce2.row, column=6).value,"已更新账号有效期至",i_list[3].strip())
                    wb_sheet2.cell(row=ce2.row, column=9).value = i_list[1].strip()
                    wb_sheet2.cell(row=ce2.row, column=10).value = i_list[2].strip()
                    wb_sheet2.cell(row=ce2.row, column=11).value = i_list[3].strip()
        if True not in num_list2:
            print(i_list[0],"跳板机账号不存在")

wb.save(file_dir)